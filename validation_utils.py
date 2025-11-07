import pandas as pd
import numpy as np

def validate_columns(df, required_columns, context=""):
    missing = [col for col in required_columns if col not in df.columns]
    if missing:
        raise ValueError(
            f"[validation_utils.py] Line 5: Missing required columns {missing} in {context}. "
            f"Columns present: {list(df.columns)}. Please ensure your file includes all required columns."
        )

def validate_no_nans(df, columns, context=""):
    for col in columns:
        nan_mask = df[col].isnull()
        if nan_mask.any():
            nan_indices = df.index[nan_mask].tolist()
            nan_preview = df.loc[nan_mask, col].head(5).tolist()
            details = [
                f"Row {idx}, Column '{col}', Value: {repr(val)}"
                for idx, val in zip(nan_indices[:5], nan_preview)
            ]
            raise ValueError(
                f"NaN (missing) values found in column '{col}' in {context}.\n"
                f"First 5 occurrences:\n" +
                "\n".join(details) +
                "\nPlease check your file for empty or invalid cells in these locations."
            )

def validate_positive_values(df, columns, context=""):
    for col in columns:
        neg_mask = df[col] < 0
        if neg_mask.any():
            neg_indices = df.index[neg_mask].tolist()
            neg_preview = df.loc[neg_mask, col].head(5).tolist()
            details = [
                f"Row {idx}, Column '{col}', Value: {val}"
                for idx, val in zip(neg_indices[:5], neg_preview)
            ]
            raise ValueError(
                f"Negative values found in column '{col}' in {context}.\n"
                f"First 5 occurrences:\n" +
                "\n".join(details) +
                "\nAll values in this column must be positive."
            )

def validate_percentage_sum(df, percentage_col, expected_sum=100, tolerance=0.5, context=""):
    total = df[percentage_col].sum()
    if not (expected_sum - tolerance <= total <= expected_sum + tolerance):
        preview = df[[percentage_col]].head(5).to_dict(orient="records")
        raise ValueError(
            f"Sum of column '{percentage_col}' is {total}, expected approximately {expected_sum} (±{tolerance}) in {context}.\n"
            f"First 5 values in '{percentage_col}': {preview}\n"
            "Please check if there are missing or incorrect values in this column."
        )

def validate_unique(df, columns, context=""):
    dup_mask = df.duplicated(subset=columns, keep=False)
    if dup_mask.any():
        dup_rows = df[dup_mask].head(5)
        details = dup_rows[columns].to_dict(orient="records")
        indices = dup_rows.index.tolist()
        raise ValueError(
            f"Duplicate rows found for columns {columns} in {context}.\n"
            f"First 5 duplicate row indices: {indices}\n"
            f"Duplicate values: {details}\n"
            "Please ensure these columns uniquely identify each row."
        )

def validate_file_exists(filepath):
    import os
    if not os.path.exists(filepath):
        raise FileNotFoundError(
            f"File not found: {filepath}\n"
            "Please check the file path and ensure the file exists."
        )

def validate_sheet_exists(filepath, sheet_name):
    import openpyxl
    wb = openpyxl.load_workbook(filepath, read_only=True)
    if sheet_name not in wb.sheetnames:
        raise ValueError(
            f"Sheet '{sheet_name}' not found in file '{filepath}'.\n"
            f"Available sheets: {wb.sheetnames}\n"
            "Please check the sheet name and try again."
        )

def validate_datetime_column(df, column, context=""):
    try:
        pd.to_datetime(df[column])
    except Exception as e:
        # Try to find the first problematic value
        for idx, val in df[column].items():
            try:
                pd.to_datetime(val)
            except Exception as e2:
                raise ValueError(
                    f"Invalid datetime value in column '{column}' at row {idx} in {context}.\n"
                    f"Value: {repr(val)}\nError: {e2}\n"
                    "Please correct or remove this value."
                ) from e
        # If not found, raise the original error
        raise ValueError(
            f"Invalid datetime in column '{column}' in {context}: {e}"
        )

def validate_nonempty(df, context=""):
    if df.empty:
        raise ValueError(
            f"DataFrame is empty in {context}.\n"
            "Please check that your file contains data."
        )


def validate_month(df, datetime_col, expected_month=None, context=""):
    """
    Validates which month the data belongs to, and optionally checks if all data is from the expected month.
    - df: pandas DataFrame
    - datetime_col: column name containing datetime values
    - expected_month: int (1-12), if provided, checks all data is from this month
    - context: string for error messages
    Returns: int (the month found in the data, if expected_month is not provided)
    Raises ValueError if data contains multiple months or does not match expected_month.
    """
    try:
        dt_series = pd.to_datetime(df[datetime_col])
        months = dt_series.dt.month.unique()
    except Exception as e:
        raise ValueError(f"Invalid datetime in column '{datetime_col}' in {context}: {e}")
    if len(months) > 1:
        # Show which rows have which months
        month_map = dt_series.dt.month
        details = []
        for m in months:
            indices = df.index[month_map == m].tolist()
            details.append(f"Month {m}: rows {indices[:5]}{'...' if len(indices) > 5 else ''}")
        raise ValueError(
            f"Data contains multiple months {months} in {context}.\n" +
            "\n".join(details) +
            "\nPlease ensure all data is from a single month."
        )
    if expected_month is not None and months[0] != expected_month:
        indices = df.index[dt_series.dt.month != expected_month].tolist()
        raise ValueError(
            f"Data month {months[0]} does not match expected {expected_month} in {context}.\n"
            f"First 5 mismatched row indices: {indices[:5]}\n"
            "Please ensure all data is from the expected month."
        )
    return months[0]

def validate_15min_granularity(df, datetime_col, context="", tolerance_seconds=60, strict=True):
    """
    Validates that all datetime values are at 15-minute intervals and consecutive rows are spaced by 15 minutes.
    - df: pandas DataFrame
    - datetime_col: column name containing datetime values
    - context: string for error messages
    - tolerance_seconds: allowable deviation in seconds for timestamp alignment (default 60s)
    - strict: if True, requires exact alignment; if False, allows tolerance
    Raises ValueError if any timestamp is not aligned to a 15-minute interval or intervals are not 15 minutes.
    """
    try:
        dt_series = pd.to_datetime(df[datetime_col])
    except Exception as e:
        raise ValueError(f"Invalid datetime in column '{datetime_col}' in {context}: {e}")

    # Check all timestamps are aligned to 15-minute intervals (with optional tolerance)
    aligned = []
    for ts in dt_series:
        # Find the nearest 15-min mark
        minute = ts.minute
        second = ts.second
        microsecond = ts.microsecond
        total_seconds = minute * 60 + second
        # Find how far from the nearest 15-min mark (0, 15, 30, 45)
        nearest = min([abs(total_seconds - m * 60) for m in [0, 15, 30, 45]])
        if strict:
            aligned.append((minute % 15 == 0) and (second == 0) and (microsecond == 0))
        else:
            aligned.append(nearest <= tolerance_seconds and microsecond == 0)
    not_15min = ~pd.Series(aligned)
    if not_15min.any():
        bad_times = dt_series[not_15min].head(5).tolist()
        bad_indices = df.index[not_15min].tolist()[:5]
        details = []
        for idx, val in zip(bad_indices, bad_times):
            minute = val.minute
            second = val.second
            total_seconds = minute * 60 + second
            nearest = min([abs(total_seconds - m * 60) for m in [0, 15, 30, 45]])
            details.append(f"Row {idx}: {val} (off by {nearest} seconds from nearest 15-min mark)")
        raise ValueError(
            f"Found timestamps not aligned to 15-minute intervals in {context}.\n" +
            "\n".join(details) +
            (f"\nAllowed tolerance: {tolerance_seconds} seconds." if not strict else "\nStrict mode: no tolerance.") +
            "\nPlease ensure all timestamps are at 00, 15, 30, or 45 minutes past the hour."
        )

    # Check consecutive intervals are 15 minutes (with optional tolerance)
    sorted_dt = dt_series.sort_values()
    diffs = sorted_dt.diff().dropna()
    if strict:
        not_15min_diff = ~(diffs == pd.Timedelta(minutes=15))
    else:
        not_15min_diff = ~(diffs.apply(lambda x: abs(x.total_seconds() - 900) <= tolerance_seconds))
    if not_15min_diff.any():
        bad_indices = diffs[not_15min_diff].index[:5].tolist()
        bad_deltas = diffs[not_15min_diff].head(5).tolist()
        details = [f"Row {idx}: interval {delta}" for idx, delta in zip(bad_indices, bad_deltas)]
        raise ValueError(
            f"Found non-15-minute intervals between consecutive timestamps in {context}.\n" +
            "\n".join(details) +
            (f"\nAllowed tolerance: {tolerance_seconds} seconds." if not strict else "\nStrict mode: no tolerance.") +
            "\nPlease ensure all consecutive timestamps are exactly 15 minutes apart."
        )

#add validation which month data
#check all the values are 15 minutes granularity
